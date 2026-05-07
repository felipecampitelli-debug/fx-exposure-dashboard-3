import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import plotly.graph_objects as go
import plotly.express as px
from io import BytesIO
import re

# ============================================================================
# CONFIGURACIÓN DE PÁGINA
# ============================================================================
st.set_page_config(
    page_title="FX Exposure Dashboard",
    page_icon="💱",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================================
# ESTILOS CSS
# ============================================================================
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: 700;
        color: #1f77b4;
        text-align: center;
        padding: 1rem 0;
        border-bottom: 3px solid #1f77b4;
        margin-bottom: 2rem;
    }
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1.5rem;
        border-radius: 10px;
        color: white;
        text-align: center;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    .metric-value {
        font-size: 2rem;
        font-weight: 700;
    }
    .metric-label {
        font-size: 0.9rem;
        opacity: 0.9;
    }
    .status-positive {
        background-color: #d4edda;
        color: #155724;
        padding: 0.25rem 0.5rem;
        border-radius: 4px;
        font-weight: 600;
    }
    .status-negative {
        background-color: #f8d7da;
        color: #721c24;
        padding: 0.25rem 0.5rem;
        border-radius: 4px;
        font-weight: 600;
    }
</style>
""", unsafe_allow_html=True)

# ============================================================================
# CONSTANTES
# ============================================================================
MONEDAS = ['ARS', 'BRL', 'MXN', 'CLP', 'COP', 'PEN', 'UYU']
NOMBRES_MONEDA = {
    'ARS': 'ARGENTINIAN PESO',
    'BRL': 'BRAZILIAN REAL',
    'MXN': 'MEXICAN PESO',
    'CLP': 'CHILEAN PESO',
    'COP': 'COLOMBIAN PESO',
    'PEN': 'PERUVIAN SOL',
    'UYU': 'URUGUAYAN PESO'
}
LIBRO_MAP = {
    'ARS':'DESPEGAR ARGENTINA CORP',
    'BRL':'DESPEGAR BRASIL CORP',
    'MXN':'DESPEGAR MEXICO OPER CORP',
    'CLP':'DESPEGAR CHILE CORP',
    'COP':'DESPEGAR COLOMBIA SAS CORP',
    'PEN':'DESPEGAR PERU CORP',
    'UYU':'DESPEGAR URUGUAY OP CORP'
}

# ============================================================================
# FUNCIONES DE PROCESAMIENTO
# ============================================================================

@st.cache_data
def procesar_datos(mayor_file, tc_file, ndf_file, ops_file, flex_file, 
                   prestamos_file, fx_exposure_file):
    """Procesa todos los archivos y retorna datos calculados"""
    
    with st.spinner('📊 Procesando datos...'):
        
        # Extraer fecha del archivo mayor
        match = re.search(r'(\d{6})\.txt$', mayor_file.name)
        if not match:
            st.error("❌ No se pudo extraer fecha del archivo mayor")
            return None
            
        dd, mm, aa = int(match.group(1)[:2]), int(match.group(1)[2:4]), int(match.group(1)[4:6])
        fecha_proceso = datetime(2000 + aa, mm, dd)
        
        # 1. CARGAR MAYOR
        df = pd.read_csv(mayor_file, sep='|', encoding='latin-1', low_memory=False)
        df['SALDO_FINAL_MON_INF'] = pd.to_numeric(df['SALDO_FINAL_MON_INF'], errors='coerce').fillna(0)
        for col in ['LIBRO', 'RUBRO', 'NOTA', 'MONEDA_ORIGEN', 'CLASIFICACION_CONTABLE', 'COMBINACION_CONTABLE']:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()
        if 'CUENTA' in df.columns:
            df['CUENTA'] = pd.to_numeric(df['CUENTA'], errors='coerce')
        
        # 2. CARGAR TCs
        df_tc = pd.read_excel(tc_file, sheet_name='CCY Spot LatAm', skiprows=1, header=None)
        df_tc[0] = pd.to_datetime(df_tc[0], errors='coerce')
        fila_tc = df_tc[df_tc[0] == fecha_proceso]
        if fila_tc.empty:
            fila_tc = df_tc[df_tc[0] <= fecha_proceso].iloc[[-1]]
        
        tc = {
            'ARS': float(fila_tc.iloc[0, 1]),
            'BRL': float(fila_tc.iloc[0, 5]),
            'MXN': float(fila_tc.iloc[0, 9]),
            'CLP': float(fila_tc.iloc[0, 13]),
            'COP': float(fila_tc.iloc[0, 17]),
            'PEN': float(fila_tc.iloc[0, 21]),
            'UYU': float(fila_tc.iloc[0, 29]),
            'USD': 1.0
        }
        
        df['TC'] = df['MONEDA_ORIGEN'].map(tc)
        df['SALDO_USD'] = df['SALDO_FINAL_MON_INF'] / df['TC']
        
        # 3. FLEX TAXES
        flex_taxes_set = set(pd.read_excel(flex_file, sheet_name='Flex Taxes')['COMBINACION_CONTABLE'].astype(str).str.strip().unique())
        
        # 4. OPS NO CAPTURADAS
        ops_no_cap = {}
        for moneda in MONEDAS:
            try:
                df_ops = pd.read_excel(ops_file, sheet_name=moneda, header=4)
                fechas_dict = {}
                for col in df_ops.columns[2:]:
                    try:
                        fecha_col = pd.to_datetime(col, errors='coerce', dayfirst=True)
                        if pd.notna(fecha_col):
                            fechas_dict[col] = fecha_col
                    except:
                        continue
                col_sel = next((c for c, f in fechas_dict.items() if f.date() == fecha_proceso.date()), None)
                if not col_sel:
                    fechas_ant = {c: f for c, f in fechas_dict.items() if f <= fecha_proceso}
                    if fechas_ant:
                        col_sel = max(fechas_ant, key=fechas_ant.get)
                if col_sel:
                    ma = pd.to_numeric(df_ops[df_ops.iloc[:, 0].astype(str).str.contains("M&A", case=False, na=False)][col_sel].values[0] if len(df_ops[df_ops.iloc[:, 0].astype(str).str.contains("M&A", case=False, na=False)]) > 0 else 0, errors='coerce')
                    fx = pd.to_numeric(df_ops[df_ops.iloc[:, 0].astype(str).str.contains("FX Spot", case=False, na=False)][col_sel].values[0] if len(df_ops[df_ops.iloc[:, 0].astype(str).str.contains("FX Spot", case=False, na=False)]) > 0 else 0, errors='coerce')
                    sp = pd.to_numeric(df_ops[df_ops.iloc[:, 0].astype(str).str.contains("Specific", case=False, na=False)][col_sel].values[0] if len(df_ops[df_ops.iloc[:, 0].astype(str).str.contains("Specific", case=False, na=False)]) > 0 else 0, errors='coerce')
                    ops_no_cap[moneda] = {'ma': 0 if pd.isna(ma) else ma, 'fx_spot': 0 if pd.isna(fx) else fx, 'specific': 0 if pd.isna(sp) else sp}
                else:
                    ops_no_cap[moneda] = {'ma': 0, 'fx_spot': 0, 'specific': 0}
            except:
                ops_no_cap[moneda] = {'ma': 0, 'fx_spot': 0, 'specific': 0}
        
        # 5. NDF
        df_ndf = pd.read_excel(ndf_file, sheet_name='Aux')
        df_ndf = df_ndf[df_ndf['Type'] != 0]
        df_ndf['Trade'] = pd.to_datetime(df_ndf['Trade'], errors='coerce')
        df_ndf['Fixing Date'] = pd.to_datetime(df_ndf['Fixing Date'], errors='coerce')
        df_ndf = df_ndf.dropna(subset=['Trade', 'Fixing Date'])
        df_ndf = df_ndf[(df_ndf['Trade'] < fecha_proceso) & (df_ndf['Fixing Date'] > fecha_proceso)]
        df_ndf['USD_Reporte'] = -df_ndf['USD']
        
        ndf_pos = df_ndf.groupby('Currency')['USD_Reporte'].sum().to_dict()
        for m in MONEDAS:
            if m not in ndf_pos:
                ndf_pos[m] = 0
        
        # Ventanas de tiempo para NDFs
        dia_semana = fecha_proceso.weekday()
        if dia_semana == 4:
            inicio_t0 = fecha_proceso + timedelta(days=3)
        elif dia_semana in [0, 1]:
            inicio_t0 = fecha_proceso + timedelta(days=-dia_semana)
        else:
            inicio_t0 = fecha_proceso + timedelta(days=7-dia_semana if dia_semana < 5 else 1)
        fin_t0 = inicio_t0 + timedelta(days=4)
        inicio_t1, fin_t1 = inicio_t0 + timedelta(days=7), fin_t0 + timedelta(days=7)
        inicio_t2, fin_t2 = inicio_t1 + timedelta(days=7), fin_t1 + timedelta(days=7)
        
        # NDF Semana 0
        df_ndf_t0 = df_ndf[(df_ndf['Fixing Date'] >= inicio_t0) & (df_ndf['Fixing Date'] <= fin_t0)]
        ndf_vencimientos_t0 = df_ndf_t0.groupby('Currency')['USD_Reporte'].sum().to_dict()
        for m in MONEDAS:
            if m not in ndf_vencimientos_t0:
                ndf_vencimientos_t0[m] = 0
        
        ndf_pl_t0 = {}
        for m in MONEDAS:
            df_m = df_ndf_t0[df_ndf_t0['Currency'] == m]
            if len(df_m) > 0:
                df_m = df_m.copy()
                df_m['P&L_USD'] = ((df_m['NDF'] - tc.get(m, 1.0)) * df_m['USD_Reporte']) / tc.get(m, 1.0)
                ndf_pl_t0[m] = df_m['P&L_USD'].sum()
            else:
                ndf_pl_t0[m] = 0
        
        # NDF Semanas 1+2
        df_ndf_t1_t2 = df_ndf[(df_ndf['Fixing Date'] >= inicio_t1) & (df_ndf['Fixing Date'] <= fin_t2)]
        ndf_vencimientos_t1_t2 = df_ndf_t1_t2.groupby('Currency')['USD_Reporte'].sum().to_dict()
        for m in MONEDAS:
            if m not in ndf_vencimientos_t1_t2:
                ndf_vencimientos_t1_t2[m] = 0
        
        ndf_pl_t1_t2 = {}
        for m in MONEDAS:
            df_m = df_ndf_t1_t2[df_ndf_t1_t2['Currency'] == m]
            if len(df_m) > 0:
                df_m = df_m.copy()
                df_m['P&L_USD'] = ((df_m['NDF'] - tc.get(m, 1.0)) * df_m['USD_Reporte']) / tc.get(m, 1.0)
                ndf_pl_t1_t2[m] = df_m['P&L_USD'].sum()
            else:
                ndf_pl_t1_t2[m] = 0
        
        # 6. LOAN CLP
        df_prest = pd.read_excel(prestamos_file, sheet_name='Viejo')
        df_prest_filt = df_prest[df_prest['Date (Cash Flow)'] <= fecha_proceso].copy()
        df_ultima = df_prest_filt.sort_values('Date (Cash Flow)').groupby('Name/Code').last().reset_index()
        col_balance = df_prest.columns[7]
        df_ultima[col_balance] = pd.to_numeric(df_ultima[col_balance], errors='coerce')
        df_validos = df_ultima[df_ultima[col_balance] > 0]
        if len(df_validos) > 0:
            df_validos = df_validos.copy()
            df_validos['USD'] = df_validos[col_balance] / tc['CLP']
            loan_position_clp = -1 * df_validos['USD'].sum()
        else:
            loan_position_clp = 0
        
        # 7. CALCULAR 69 LÍNEAS POR MONEDA
        datos_moneda = {}
        pivot_total = df.groupby(['LIBRO', 'MONEDA_ORIGEN'])[['SALDO_FINAL_MON_INF']].sum().reset_index()
        
        for moneda in MONEDAS:
            L = {}
            L['L3'] = df[(df['MONEDA_ORIGEN']==moneda) & (df['RUBRO']=='Cash and cash equivalents')]['SALDO_USD'].sum()
            L['L4'] = df[(df['MONEDA_ORIGEN']==moneda) & (df['RUBRO']=='Accounts receivable, net of allowances') & (df['NOTA']!='Credit Card vouchers D!')]['SALDO_USD'].sum()
            L['L5'] = df[(df['MONEDA_ORIGEN']==moneda) & (df['RUBRO']=='Restricted Cash')]['SALDO_USD'].sum()
            L['L6'] = df[(df['MONEDA_ORIGEN']==moneda) & (df['RUBRO']=='Other receivables and prepaid expenses') & (df['NOTA']!='Tax Credits')]['SALDO_USD'].sum()
            L['L7'] = df[(df['MONEDA_ORIGEN']==moneda) & (df['RUBRO']=='Related party receivable')]['SALDO_USD'].sum()
            L['L8'] = df[(df['MONEDA_ORIGEN']==moneda) & (df['NOTA']=='Tax Credits') & (df['COMBINACION_CONTABLE'].isin(flex_taxes_set)) & (df['CLASIFICACION_CONTABLE']=='Current Assets')]['SALDO_USD'].sum()
            L['L2'] = L['L3']+L['L4']+L['L5']+L['L6']+L['L7']+L['L8']
            L['L12'] = df[(df['MONEDA_ORIGEN']==moneda) & (df['RUBRO']=='Tourism Suppliers Payable') & (df['NOTA']=='Airlines')]['SALDO_USD'].sum()
            L['L13'] = df[(df['MONEDA_ORIGEN']==moneda) & (df['RUBRO']=='Tourism Suppliers Payable') & (df['NOTA']=='Hotel and ONA Suppliers') & (~df['CUENTA'].isin([21211,21261]))]['SALDO_USD'].sum()
            L['L14'] = df[(df['MONEDA_ORIGEN']==moneda) & (df['RUBRO']=='Tourism Suppliers Payable') & (df['NOTA']=='Other tourism suppliers')]['SALDO_USD'].sum()
            L['L11'] = L['L12']+L['L13']+L['L14']
            L['L15'] = df[(df['MONEDA_ORIGEN']==moneda) & (df['RUBRO']=='Accounts payable')]['SALDO_USD'].sum()
            L['L16'] = df[(df['MONEDA_ORIGEN']==moneda) & (df['RUBRO']=='Lease liabilities') & (df['CLASIFICACION_CONTABLE']=='Current Liabilities')]['SALDO_USD'].sum()
            L['L17'] = df[(df['MONEDA_ORIGEN']==moneda) & (df['RUBRO']=='Salaries and social security payable')]['SALDO_USD'].sum()
            L['L18'] = df[(df['MONEDA_ORIGEN']==moneda) & (df['RUBRO']=='Related party payable')]['SALDO_USD'].sum()
            L['L19'] = df[(df['MONEDA_ORIGEN']==moneda) & (df['RUBRO']=='Other liabilities') & (df['CLASIFICACION_CONTABLE']=='Current Liabilities')]['SALDO_USD'].sum()
            L['L20'] = df[(df['MONEDA_ORIGEN']==moneda) & (df['NOTA']=='Taxes payable') & (df['COMBINACION_CONTABLE'].isin(flex_taxes_set)) & (df['CLASIFICACION_CONTABLE']=='Current Liabilities')]['SALDO_USD'].sum()
            L['L10'] = L['L11']+L['L15']+L['L16']+L['L17']+L['L18']+L['L19']+L['L20']
            L['L22'] = L['L2']+L['L10']
            L['L24'] = ndf_pos.get(moneda,0)
            L['L25'] = loan_position_clp if moneda=='CLP' else 0
            L['L26'] = ops_no_cap[moneda]['ma']
            L['L27'] = ops_no_cap[moneda]['fx_spot']
            L['L28'] = ops_no_cap[moneda]['specific']
            L['L30'] = L['L22']+L['L24']+L['L25']+L['L26']+L['L27']+L['L28']
            datos_moneda[moneda] = L
        
        return {
            'fecha_proceso': fecha_proceso,
            'tc': tc,
            'datos_moneda': datos_moneda,
            'ndf_pos': ndf_pos,
            'ndf_vencimientos_t0': ndf_vencimientos_t0,
            'ndf_pl_t0': ndf_pl_t0,
            'ndf_vencimientos_t1_t2': ndf_vencimientos_t1_t2,
            'ndf_pl_t1_t2': ndf_pl_t1_t2,
            'loan_position_clp': loan_position_clp,
            'inicio_t0': inicio_t0,
            'fin_t0': fin_t0,
            'inicio_t1': inicio_t1,
            'fin_t2': fin_t2
        }

def generar_excel(datos):
    """Genera el archivo Excel completo"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # Hoja FX
    ws_fx = wb.create_sheet('FX', 0)
    ws_fx['A1'] = 'FX EXPOSURE REPORT'
    ws_fx['A1'].font = Font(size=14, bold=True)
    ws_fx['A2'] = datos['fecha_proceso'].strftime('%d-%b-%Y')
    ws_fx['A4'], ws_fx['B4'] = 'Moneda', 'TC'
    for cell in ['A4', 'B4']:
        ws_fx[cell].font = Font(bold=True)
        ws_fx[cell].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    for i, m in enumerate(MONEDAS, start=5):
        ws_fx[f'A{i}'], ws_fx[f'B{i}'] = m, datos['tc'][m]
        ws_fx[f'B{i}'].number_format = '0.0000'
    
    # Hojas por moneda (simplificado)
    for idx, m in enumerate(MONEDAS, start=2):
        ws = wb.create_sheet(m, idx)
        ws['A1'] = NOMBRES_MONEDA[m]
        ws['A1'].font = Font(bold=True, size=11, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='42426C', end_color='42426C', fill_type='solid')
        ws['A3'] = 'Working Capital Exposure'
        ws['B3'] = datos['datos_moneda'][m]['L30'] / 1_000_000
        ws['B3'].number_format = '$#,##0.0'
    
    # Guardar en BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ============================================================================
# INTERFAZ STREAMLIT
# ============================================================================

st.markdown('<div class="main-header">💱 FX Exposure Dashboard</div>', unsafe_allow_html=True)

# ============================================================================
# SIDEBAR - CARGA DE ARCHIVOS
# ============================================================================

with st.sidebar:
    st.header("📁 Cargar Archivos")
    st.markdown("---")
    
    mayor_file = st.file_uploader(
        "1️⃣ Mayor Contable",
        type=['txt'],
        help="gl_balance_de_comprobacion_DDMMAA.txt"
    )
    
    tc_file = st.file_uploader(
        "2️⃣ Tipos de Cambio",
        type=['xlsx'],
        help="ccy_spot_latam.xlsx"
    )
    
    ndf_file = st.file_uploader(
        "3️⃣ Posiciones NDF",
        type=['xlsx', 'xlsm'],
        help="ndf_position*.xlsx"
    )
    
    ops_file = st.file_uploader(
        "4️⃣ Ops No Capturadas",
        type=['xlsx'],
        help="ops_no_capturadas*.xlsx"
    )
    
    flex_file = st.file_uploader(
        "5️⃣ Flex Taxes",
        type=['xlsx'],
        help="flex_taxes.xlsx"
    )
    
    prestamos_file = st.file_uploader(
        "6️⃣ Préstamos Chile",
        type=['xlsx'],
        help="vtos_prestamos_chile.xlsx"
    )
    
    fx_exposure_file = st.file_uploader(
        "7️⃣ FX Exposure Histórico",
        type=['xlsx'],
        help="fx_exposure_YYYYMMDD.xlsx"
    )
    
    st.markdown("---")
    
    archivos_ok = all([mayor_file, tc_file, ndf_file, ops_file, flex_file, 
                       prestamos_file, fx_exposure_file])
    
    if archivos_ok:
        st.success("✅ Todos los archivos cargados")
        procesar_btn = st.button("🚀 GENERAR ANÁLISIS", type="primary", use_container_width=True)
    else:
        st.warning(f"⚠️ Faltan {7 - sum([bool(f) for f in [mayor_file, tc_file, ndf_file, ops_file, flex_file, prestamos_file, fx_exposure_file]])} archivos")
        procesar_btn = False

# ============================================================================
# CONTENIDO PRINCIPAL
# ============================================================================

if not archivos_ok:
    st.info("👈 Por favor carga los 7 archivos en el panel lateral para comenzar")
    
    # Mostrar instrucciones
    st.markdown("### 📋 Instrucciones")
    st.markdown("""
    1. **Carga los archivos** requeridos en el panel lateral
    2. **Presiona el botón** "GENERAR ANÁLISIS"
    3. **Visualiza** las métricas de exposición
    4. **Descarga** el Excel generado
    
    **Archivos requeridos:**
    - Mayor contable (`.txt`)
    - Tipos de cambio (`.xlsx`)
    - Posiciones NDF (`.xlsx` o `.xlsm`)
    - Operaciones no capturadas (`.xlsx`)
    - Flex Taxes (`.xlsx`)
    - Préstamos Chile (`.xlsx`)
    - FX Exposure histórico (`.xlsx`)
    """)

elif procesar_btn:
    
    # Procesar datos
    datos = procesar_datos(mayor_file, tc_file, ndf_file, ops_file, 
                          flex_file, prestamos_file, fx_exposure_file)
    
    if datos is None:
        st.error("❌ Error al procesar los datos")
        st.stop()
    
    # Guardar en session_state
    st.session_state['datos'] = datos
    st.session_state['procesado'] = True
    
    st.success(f"✅ Análisis generado para {datos['fecha_proceso'].strftime('%d-%b-%Y')}")

# Mostrar dashboard si ya está procesado
if st.session_state.get('procesado', False):
    
    datos = st.session_state['datos']
    
    # ========================================================================
    # KPIs PRINCIPALES
    # ========================================================================
    
    st.markdown("### 📊 Métricas Principales")
    
    total_wce = sum(datos['datos_moneda'][m]['L30'] for m in MONEDAS) / 1e6
    total_ndf = sum(abs(datos['ndf_pos'].get(m, 0)) for m in MONEDAS) / 1e6
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Working Capital Exposure</div>
            <div class="metric-value">${total_wce:,.1f}M</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Total NDF Position</div>
            <div class="metric-value">${total_ndf:,.1f}M</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        exposure_after = total_wce + total_ndf
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Exposure After NDFs</div>
            <div class="metric-value">${exposure_after:,.1f}M</div>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # ========================================================================
    # TABLA POR MONEDA
    # ========================================================================
    
    st.markdown("### 🌎 Exposición por Moneda")
    
    tabla_data = []
    for m in MONEDAS:
        wce = datos['datos_moneda'][m]['L30'] / 1e6
        ndf = datos['ndf_pos'].get(m, 0) / 1e6
        after = wce + ndf
        tabla_data.append({
            'Moneda': m,
            'WC Exposure (MM)': f"${wce:,.1f}",
            'NDF Position (MM)': f"${ndf:,.1f}",
            'After NDFs (MM)': f"${after:,.1f}",
            'TC': f"{datos['tc'][m]:,.4f}"
        })
    
    df_tabla = pd.DataFrame(tabla_data)
    st.dataframe(df_tabla, use_container_width=True, hide_index=True)
    
    st.markdown("---")
    
    # ========================================================================
    # GRÁFICOS
    # ========================================================================
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("#### 📈 WC Exposure por Moneda")
        fig1 = go.Figure(data=[
            go.Bar(
                x=MONEDAS,
                y=[datos['datos_moneda'][m]['L30'] / 1e6 for m in MONEDAS],
                marker_color=['#1f77b4' if datos['datos_moneda'][m]['L30'] > 0 else '#d62728' for m in MONEDAS],
                text=[f"${datos['datos_moneda'][m]['L30'] / 1e6:,.1f}M" for m in MONEDAS],
                textposition='outside'
            )
        ])
        fig1.update_layout(
            height=400,
            showlegend=False,
            yaxis_title="USD Millones",
            xaxis_title="Moneda"
        )
        st.plotly_chart(fig1, use_container_width=True)
    
    with col2:
        st.markdown("#### 💼 NDF Position por Moneda")
        fig2 = go.Figure(data=[
            go.Bar(
                x=MONEDAS,
                y=[datos['ndf_pos'].get(m, 0) / 1e6 for m in MONEDAS],
                marker_color='#2ca02c',
                text=[f"${datos['ndf_pos'].get(m, 0) / 1e6:,.1f}M" for m in MONEDAS],
                textposition='outside'
            )
        ])
        fig2.update_layout(
            height=400,
            showlegend=False,
            yaxis_title="USD Millones",
            xaxis_title="Moneda"
        )
        st.plotly_chart(fig2, use_container_width=True)
    
    st.markdown("---")
    
    # ========================================================================
    # NDFs VENCIMIENTOS
    # ========================================================================
    
    st.markdown("### 📅 Cronograma de NDFs")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown(f"#### Semana Actual ({datos['inicio_t0'].strftime('%d-%b')} al {datos['fin_t0'].strftime('%d-%b')})")
        
        ndf_t0_data = []
        for m in MONEDAS:
            notional = datos['ndf_vencimientos_t0'].get(m, 0) / 1e6
            pl = datos['ndf_pl_t0'].get(m, 0) / 1e6
            if abs(notional) > 0.1:
                ndf_t0_data.append({
                    'CCY': m,
                    'Notional (MM)': f"${abs(notional):,.1f}",
                    'P&L (MM)': f"${pl:,.2f}"
                })
        
        if ndf_t0_data:
            st.dataframe(pd.DataFrame(ndf_t0_data), use_container_width=True, hide_index=True)
        else:
            st.info("Sin vencimientos esta semana")
    
    with col2:
        st.markdown(f"#### Próximas 2 Semanas ({datos['inicio_t1'].strftime('%d-%b')} al {datos['fin_t2'].strftime('%d-%b')})")
        
        ndf_t1_t2_data = []
        for m in MONEDAS:
            notional = datos['ndf_vencimientos_t1_t2'].get(m, 0) / 1e6
            pl = datos['ndf_pl_t1_t2'].get(m, 0) / 1e6
            if abs(notional) > 0.1:
                ndf_t1_t2_data.append({
                    'CCY': m,
                    'Notional (MM)': f"${abs(notional):,.1f}",
                    'P&L (MM)': f"${pl:,.2f}"
                })
        
        if ndf_t1_t2_data:
            st.dataframe(pd.DataFrame(ndf_t1_t2_data), use_container_width=True, hide_index=True)
        else:
            st.info("Sin vencimientos próximas 2 semanas")
    
    st.markdown("---")
    
    # ========================================================================
    # DESCARGA EXCEL
    # ========================================================================
    
    st.markdown("### 📥 Descargar Reporte")
    
    excel_buffer = generar_excel(datos)
    
    st.download_button(
        label="📊 Descargar Excel Completo",
        data=excel_buffer,
        file_name=f"FX_Exposure_{datos['fecha_proceso'].strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True
    )
    
    st.markdown("---")
    st.caption(f"📅 Reporte generado: {datos['fecha_proceso'].strftime('%d-%b-%Y')}")

